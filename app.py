import streamlit as st
import pandas as pd
import os
from io import BytesIO


st.set_page_config(page_title="Mpesa Reconciliation App", layout="wide")
st.title("📊 Mpesa Reconciliation App")

uploaded_key = st.file_uploader("🔑 Upload Key File", type=["csv", "xlsx"])
uploaded_aspire = st.file_uploader("📁 Upload Aspire File", type=["csv", "xlsx"])
uploaded_safaricom = st.file_uploader("📁 Upload Safaricom File", type=["csv", "xlsx"])

if uploaded_key and uploaded_aspire and uploaded_safaricom:
    st.success("✅ All files uploaded. Click the button below to start processing.")

    process = st.button("▶️ Start Processing")

    if process:
        # Load files from uploads only
        key = pd.read_csv(uploaded_key) if uploaded_key.name.endswith('.csv') else pd.read_excel(uploaded_key)
        aspire = pd.read_csv(uploaded_aspire) if uploaded_aspire.name.endswith('.csv') else pd.read_excel(uploaded_aspire)
        safaricom = pd.read_csv(uploaded_safaricom) if uploaded_safaricom.name.endswith('.csv') else pd.read_excel(uploaded_safaricom)

        st.subheader("🔍 Key File Preview")
        st.dataframe(key.head())
        st.subheader("🧾 Aspire File Preview")
        st.dataframe(aspire.head())
        st.subheader("📨 Safaricom File Preview")
        st.dataframe(safaricom.head())

       #!/usr/bin/env python
       # coding: utf-8

       # #1.Import Relevant Libraries and Files.

       # In[ ]:


       import pandas as pd
       import openpyxl
       from openpyxl import Workbook
       from openpyxl.styles import Font
       from openpyxl.utils import get_column_letter
       # [REMOVED FOR DEPLOYMENT] from google.colab import files
       import re
       from datetime import datetime


       # In[ ]:


       # Load CSV and Excel files
       # [REMOVED FOR DEPLOYMENT] aspire = pd.read_csv('/content/MPESA_TRNS_2025-06-11.csv')
       # [REMOVED FOR DEPLOYMENT] safaricom = pd.read_csv('/content/Mpesa_852182_20250611235959.csv')
       # [REMOVED FOR DEPLOYMENT] key = pd.read_excel('/content/key.xlsx')
       # [REMOVED FOR DEPLOYMENT] key = pd.read_excel('/content/key.xlsx')

       # Preview
       aspire.head()


       # ##Align data headings to the data

       # In[ ]:


       # Step 1: Shift column names one position to the left
       new_columns = safaricom.columns[1:].tolist() + ['EXTRA']

       # Step 2: Apply the new column names
       safaricom.columns = new_columns

       # Step 3: Drop the now-unwanted 'EXTRA' column
       safaricom = safaricom.drop(columns='EXTRA')

       # Optional: Check result
       safaricom.head()


       # 

       # ##Retain relevant columns for analysis and reprts

       # In[ ]:


       # ✅ Safely select only the columns that exist in safaricom
       safaricom_cols = [
           'STORE_NAME', 'RECEIPT_NUMBER', 'ACCOUNT_TYPE_NAME', 'TRANSACTION_TYPE', 'START_TIMESTAMP',
           'TRANSACTION_PARTY_DETAILS', 'CREDIT_AMOUNT', 'DEBIT_AMOUNT', 'BALANCE', 'LINKED_TRANSACTION_ID'
       ]

       # Only select columns that actually exist to avoid KeyError
       safaricom = safaricom[[col for col in safaricom_cols if col in safaricom.columns]]

       # Preview
       safaricom.head()



       # ##Filter out Columns not in use -

       # In[ ]:


       aspire.head()


       # ##Append new columns to safaricom for reconciliation

       # In[ ]:


       key.columns = ['Original_STORE_NAME', 'Clean_STORE_NAME']


       # ##Map the Safaricom data to Aspire data

       # In[ ]:


       key.columns = ['Original_STORE_NAME', 'Clean_STORE_NAME']

       # Create mapping dictionary
       store_map = dict(zip(key['Original_STORE_NAME'], key['Clean_STORE_NAME']))

       # Apply to safaricom
       safaricom['Store_amend'] = safaricom['STORE_NAME'].map(store_map)


       # ##Check validity and duplication of the Aspire Mpesa codes

       # In[ ]:


       # Step 1: Create a set of all valid transaction IDs for faster lookup
       valid_transaction_ids = set(aspire['TRANSACTION_ID'])

       # Step 2: Apply logic to create Code_check column
       safaricom['code_check'] = safaricom['RECEIPT_NUMBER'].apply(
           lambda x: x if x in valid_transaction_ids else 'XX'
       )

       # Step 3: Remove duplicate 'code_check' column if any exist (e.g., from merge)
       safaricom = safaricom.loc[:, ~safaricom.columns.duplicated()]

       # Step 4: Count how many were marked as 'XX'
       xx_count = (safaricom['code_check'] == 'XX').sum()

       # Step 5: Display result
       st.write(f"Number of unmatched RECEIPT_NUMBERs marked as 'XX': {xx_count}")
       safaricom.head()


       # ###*Safaricom Mpesa data test Script*

       # In[ ]:


       safaricom.iloc[[109177]]


       # In[ ]:


       # Ensure TRANSACTION_ID is string
       safaricom['RECEIPT_NUMBER'] = safaricom['RECEIPT_NUMBER'].astype(str)

       # Extract the first 3 characters and create Ref1 column
       safaricom['Ref1'] = safaricom['RECEIPT_NUMBER'].str[:3]
       safaricom.head()


       # ##Append new Columns in Aspire Data for Reconcilliations

       # In[ ]:


       # Ensure TRANSACTION_ID is string
       aspire['TRANSACTION_ID'] = aspire['TRANSACTION_ID'].astype(str)

       # Extract the first 3 characters and create Ref1 column
       aspire['Ref1'] = aspire['TRANSACTION_ID'].str[:3]
       aspire.head()



       # In[ ]:


       # Create a set of valid receipt numbers from safaricom
       valid_receipts = set(safaricom['RECEIPT_NUMBER'])

       # Check existence and assign 'Yes' or 'No'
       aspire['CODE_VALIDATION'] = aspire['TRANSACTION_ID'].apply(
           lambda x: 'Yes' if x in valid_receipts else 'No'
       )
       aspire.tail()


       # ##Match Store names

       # In[ ]:


       aspire['STORE_NAME'] = aspire['STORE_NAME'].replace('OUTERING 2', 'Outering 2')


       # In[ ]:


       # Step 1: Clean up and prepare
       aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
       preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 2: Identify duplicate TRANSACTION_IDs
       duplicate_ids = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]

       # Step 3: Prioritize valid transaction types
       duplicate_ids['type_priority'] = duplicate_ids['TRANSACTION_TYPE'].apply(
           lambda x: 0 if x in preferred_types else (1 if x else 2)
       )

       # Step 4: Keep the best row per duplicate group
       cleaned_duplicates = (
           duplicate_ids.sort_values(by='type_priority')
                        .drop_duplicates(subset='TRANSACTION_ID', keep='first')
                        .drop(columns='type_priority')
       )

       # Step 5: Drop all rows with duplicated TRANSACTION_IDs from original
       aspire_nondupes = aspire[~aspire['TRANSACTION_ID'].isin(duplicate_ids['TRANSACTION_ID'])]

       # Step 6: Combine cleaned duplicates with original unique rows
       aspire = pd.concat([aspire_nondupes, cleaned_duplicates], ignore_index=True)

       # Optional: Reset index and preview
       aspire = aspire.sort_values(by='TRANSACTION_ID').reset_index(drop=True)
       aspire.head()


       # In[ ]:


       # View rows with duplicated TRANSACTION_ID (including all copies)
       duplicates = aspire[aspire.duplicated(subset='TRANSACTION_ID', keep=False)]
       duplicates


       # In[ ]:


       aspire = aspire.drop_duplicates(subset='TRANSACTION_ID', keep='first')


       # ###*Get a status Count of the daily transactions*

       # In[ ]:


       code_validation_summary = aspire['CODE_VALIDATION'].value_counts().reset_index()
       code_validation_summary.columns = ['Validation_Status', 'Count']
       code_validation_summary
       # Count validation results
       code_validation_summary = aspire['CODE_VALIDATION'].value_counts().reset_index()
       code_validation_summary.columns = ['Validation_Status', 'Count']

       # Add percentage column
       total_validation = code_validation_summary['Count'].sum()
       code_validation_summary['Percentage'] = round((code_validation_summary['Count'] / total_validation) * 100, 2)

       code_validation_summary



       # In[ ]:





       # In[ ]:


       # Load key.xlsx and rename columns properly
       # [REMOVED FOR DEPLOYMENT] key = pd.read_excel('/content/key.xlsx', header=None)
       key.columns = ['TRANSACTION_TYPE', 'Summary_type']


       # ##Check validity of the transactions

       # In[ ]:


       # Load key file and drop rows with empty or invalid TRANSACTION_TYPE
       # [REMOVED FOR DEPLOYMENT] key = pd.read_excel('/content/key.xlsx', header=None)
       key.columns = ['TRANSACTION_TYPE', 'Summary_type']

       # Drop rows where TRANSACTION_TYPE is null or in the list of irrelevant values
       exclude = ['PHONE', 'CHANNEL', 'TRANSACTION_TYPE']
       key = key[~key['TRANSACTION_TYPE'].isin(exclude)]
       key = key.dropna(subset=['TRANSACTION_TYPE'])

       aspire.tail()


       # In[ ]:


       # Merge to create the Summary_type column
       aspire = aspire.merge(key, on='TRANSACTION_TYPE', how='left')


       # ###Summary of Utilization Status

       # In[ ]:


       summary_counts = aspire['Summary_type'].value_counts().reset_index()
       summary_counts.columns = ['Summary_type', 'Count']
       summary_counts
       # Count and create summary
       summary_counts = aspire['Summary_type'].value_counts().reset_index()
       summary_counts.columns = ['Summary_type', 'Count']

       # Calculate total
       total = summary_counts['Count'].sum()

       # Add percentage column
       summary_counts['Percentage'] = round((summary_counts['Count'] / total) * 100, 2)

       summary_counts



       # In[ ]:


       aspire.head()


       # #2.Generate the daily reconcilliation summary

       # ###List all stores

       # In[ ]:


       store_summary = pd.DataFrame(safaricom['Store_amend'].dropna().unique(), columns=['Store_amend']) #Create Dataframe
       store_summary = store_summary.sort_values(by='Store_amend').reset_index(drop=True) #Sort alphabeticaly
       store_summary.head()


       # ###Compute daily Charges per store

       # In[ ]:


       # Step 1: Define charge descriptions
       charge_keywords = [
           'Pay merchant Charge',
           'FSI to Merchant Charge by Receiver',
           'Merchant to Merchant Payment Charge to M-PESA'
       ]

       # Step 2: Ensure DEBIT_AMOUNT is numeric
       safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')

       # Step 3: Filter relevant charge transactions
       safaricom_charges = safaricom[safaricom['TRANSACTION_PARTY_DETAILS'].isin(charge_keywords)]

       # Step 4: Group charges by Store_amend
       charges_by_store = safaricom_charges.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
       charges_by_store.columns = ['Store_amend', 'Charges']

       # ✅ Step 4.5: Drop all existing 'Charges' columns in store_summary
       store_summary = store_summary.loc[:, ~store_summary.columns.str.contains('Charges', case=False)]

       # Step 5: Merge clean charges into store_summary1
       store_summary1 = store_summary.merge(charges_by_store, on='Store_amend', how='left')
       store_summary1['Charges'] = store_summary1['Charges'].fillna(0)

       # Step 6: Remove any existing TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']

       # Step 7: Add TOTAL row
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # Step 8: Show result with only one clean Charges column
       store_summary1.tail()


       # ###Check Previous day utilization

       # In[ ]:


       # Step 0: Clean up TRANSACTION_TYPE for prioritization
       aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
       preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 0.1: Identify duplicated TRANSACTION_IDs
       dupes = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]

       # Step 0.2: Add priority column: valid → non-empty → empty
       dupes['priority'] = dupes['TRANSACTION_TYPE'].apply(
           lambda x: 0 if x in preferred_types else (1 if x else 2)
       )

       # Step 0.3: Keep one best row per TRANSACTION_ID
       cleaned_dupes = (
           dupes.sort_values(by='priority')
                .drop_duplicates(subset='TRANSACTION_ID', keep='first')
                .drop(columns='priority')
       )

       # Step 0.4: Combine cleaned dupes with non-dupes
       aspire_unique = aspire[~aspire['TRANSACTION_ID'].isin(dupes['TRANSACTION_ID'])]
       aspire = pd.concat([aspire_unique, cleaned_dupes], ignore_index=True)

       # Step 1: Get the most frequent Ref1 in safaricom
       most_common_ref1 = safaricom['Ref1'].mode()[0]

       # Step 2: Define valid transaction types
       valid_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 3: Filter aspire for valid rows
       filtered_aspire = aspire[
           (aspire['Ref1'] != most_common_ref1) &
           (aspire['TRANSACTION_TYPE'].isin(valid_types))
       ]

       # Step 4: Group by STORE_NAME and sum AMOUNT
       prev_day_data = filtered_aspire.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
       prev_day_data.columns = ['Store_amend', 'Prev_day']  # Rename for merging

       # ✅ Step 4.5: Drop any existing 'Prev_day' column from store_summary1
       store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Prev_day', case=False)]

       # Step 5: Merge the computed Prev_day into store_summary1
       store_summary1 = store_summary1.merge(prev_day_data, on='Store_amend', how='left')
       store_summary1['Prev_day'] = store_summary1['Prev_day'].fillna(0)

       # Step 6: Remove any existing TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']

       # Step 7: Add TOTAL row
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # Step 8: Show result
       store_summary1.tail(20)


       # In[ ]:


       aspire[aspire['TRANSACTION_ID'] == 'TEJ3FKUUKN']


       # In[ ]:


       store_summary1['Store_match'] = store_summary1['Store_amend'].isin(aspire['STORE_NAME'].unique())
       store_summary1.tail(20)



       # ###Compute amount transfered to bank per store

       # In[ ]:


       # ✅ Step 1: Remove existing Bank_Transfer and Store_match columns (if they exist)
       store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Bank_Transfer|Store_match', case=False)]

       # Step 2: Ensure DEBIT_AMOUNT is numeric
       safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')

       # Step 3: Filter safaricom for valid transfer rows
       bank_transfer_data = safaricom[
           safaricom['TRANSACTION_PARTY_DETAILS'].str.contains(
               'Merchant Account to Organization Settlement Account', case=False, na=False
           )
       ]

       # Step 4: Group total transfers by Store_amend
       bank_transfer_sum = bank_transfer_data.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
       bank_transfer_sum.columns = ['Store_amend', 'Bank_Transfer']

       # Step 5: Merge Bank_Transfer into store_summary1
       store_summary1 = store_summary1.merge(bank_transfer_sum, on='Store_amend', how='left')
       store_summary1['Bank_Transfer'] = store_summary1['Bank_Transfer'].fillna(0)

       # Step 6: Move Bank_Transfer column after Store_amend
       cols = list(store_summary1.columns)
       cols.remove('Bank_Transfer')
       insert_at = cols.index('Store_amend') + 1
       cols.insert(insert_at, 'Bank_Transfer')
       store_summary1 = store_summary1[cols]

       # Step 7: Update TOTAL row (last row)
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[store_summary1['Store_amend'] != 'TOTAL'][numeric_cols].sum().to_dict()

       for col, value in totals.items():
           store_summary1.at[store_summary1.index[-1], col] = value

       # Step 8: Done. Preview last few rows
       store_summary1.tail()


       # ###Compute the day Mpesa transactions Utilized

       # In[ ]:


       # Step 1: Get the mode value of Ref1 in safaricom
       common_ref1 = safaricom['Ref1'].mode()[0]

       # Step 2: Define valid transaction summary types
       valid_summaries = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 3: Filter aspire with matching criteria
       utilized_data = aspire[
           (aspire['Ref1'] == common_ref1) &
           (aspire['Summary_type'].isin(valid_summaries))
       ]

       # Step 4: Group by STORE_NAME and sum AMOUNT
       utilized_summary = utilized_data.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
       utilized_summary.columns = ['Store_amend', 'Asp_Utilized']

       # Step 5: Remove existing 'Asp_Utilized' column if it exists
       store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Asp_Utilized', case=False)]

       # Step 6: Merge the result into store_summary1
       store_summary1 = store_summary1.merge(utilized_summary, on='Store_amend', how='left')
       store_summary1['Asp_Utilized'] = store_summary1['Asp_Utilized'].fillna(0)

       # Step 7: Reorder to place 'Asp_Utilized' after 'Prev_day'
       cols = list(store_summary1.columns)
       cols.remove('Asp_Utilized')
       insert_at = cols.index('Prev_day') + 1
       cols.insert(insert_at, 'Asp_Utilized')
       store_summary1 = store_summary1[cols]

       # Step 8: Update TOTAL row
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[store_summary1['Store_amend'] != 'TOTAL'][numeric_cols].sum().to_dict()
       for col, val in totals.items():
           store_summary1.at[store_summary1.index[-1], col] = val

       # Step 9: Show result
       store_summary1.tail()


       # ###Compute Gross payments (Utilized & Unutilized)

       # In[ ]:


       # Step 1: Ensure CREDIT_AMOUNT is numeric
       safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')

       # Step 2: Filter safaricom data
       saf_paid_data = safaricom[
           safaricom['ACCOUNT_TYPE_NAME'].str.strip().eq('Merchant Account')
       ]

       # Step 3: Group by Store_amend and sum CREDIT_AMOUNT
       saf_paid_summary = saf_paid_data.groupby('Store_amend')['CREDIT_AMOUNT'].sum().reset_index()
       saf_paid_summary.columns = ['Store_amend', 'saf_paid']

       # Step 4: Drop existing 'saf_paid' if already exists to avoid duplication
       store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('saf_paid', case=False)]

       # Step 5: Merge into store_summary1
       store_summary1 = store_summary1.merge(saf_paid_summary, on='Store_amend', how='left')
       store_summary1['saf_paid'] = store_summary1['saf_paid'].fillna(0)

       # Step 6: Reorder columns - insert 'saf_paid' after 'Asp_Utilized'
       cols = list(store_summary1.columns)
       cols.remove('saf_paid')
       insert_at = cols.index('Asp_Utilized') + 1
       cols.insert(insert_at, 'saf_paid')
       store_summary1 = store_summary1[cols]

       # Step 7: Recalculate TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Step 8: Preview result
       store_summary1.tail()


       # ###Compute Unutilized transactions

       # In[ ]:


       # Step 1: Create new column unutilized_txn
       store_summary1['unutilized_txn'] = store_summary1['saf_paid'] - store_summary1['Asp_Utilized']

       # Step 2: Move 'unutilized_txn' to appear after 'saf_paid'
       cols = list(store_summary1.columns)
       cols.remove('unutilized_txn')
       insert_at = cols.index('saf_paid') + 1
       cols.insert(insert_at, 'unutilized_txn')
       store_summary1 = store_summary1[cols]

       # Step 3: Update TOTAL row at the bottom
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Preview final result
       store_summary1.tail()


       # ###Compute Reversals done for the day

       # In[ ]:


       # Step 1: Ensure DEBIT_AMOUNT is numeric
       safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')

       # Step 2: Filter for rows where LINKED_TRANSACTION_ID is not null
       reversals_data = safaricom[
           safaricom['LINKED_TRANSACTION_ID'].notna()
       ]

       # Step 3: Group by Store_amend and sum DEBIT_AMOUNT
       reversals_sum = reversals_data.groupby('Store_amend')['DEBIT_AMOUNT'].sum().reset_index()
       reversals_sum.columns = ['Store_amend', 'Reversals']

       # Step 4: Merge into store_summary1
       store_summary1 = store_summary1.merge(reversals_sum, on='Store_amend', how='left')
       store_summary1['Reversals'] = store_summary1['Reversals'].fillna(0)

       # Step 5: Reorder column after 'unutilized_txn' if it exists
       cols = list(store_summary1.columns)
       cols.remove('Reversals')
       insert_at = cols.index('unutilized_txn') + 1 if 'unutilized_txn' in cols else len(cols)
       cols.insert(insert_at, 'Reversals')
       store_summary1 = store_summary1[cols]

       # Step 6: Update TOTAL row at the bottom
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Done: Preview result
       store_summary1.tail()


       # In[ ]:


       # Step 1: Remove all duplicate variants of Asp_Pending except the last one
       asp_cols = [col for col in store_summary1.columns if 'Asp_Pending' in col]
       if len(asp_cols) > 1:
           store_summary1['Asp_Pending'] = store_summary1[asp_cols[-1]]
           store_summary1 = store_summary1.drop(columns=[col for col in asp_cols if col != 'Asp_Pending'])

       # Step 2: Move 'Asp_Pending' next to 'Asp_Utilized'
       if 'Asp_Pending' in store_summary1.columns:
           cols = list(store_summary1.columns)
           cols.remove('Asp_Pending')
           insert_at = cols.index('Asp_Utilized') + 1 if 'Asp_Utilized' in cols else len(cols)
           cols.insert(insert_at, 'Asp_Pending')
           store_summary1 = store_summary1[cols]

       # Step 3: Remove any old TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']

       # Step 4: Recompute TOTAL row
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Final Preview
       store_summary1.tail()



       # In[ ]:


       # Step 1: Ensure AMOUNT is numeric
       aspire['AMOUNT'] = pd.to_numeric(aspire['AMOUNT'], errors='coerce')

       # Step 2: Identify most frequent Ref1 value
       most_common_ref1 = aspire['Ref1'].mode()[0]

       # Step 3: Define excluded transaction types
       excluded_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 4: Filter aspire for pending transactions
       asp_pending_data = aspire[
           (aspire['Ref1'] == most_common_ref1) &
           (~aspire['TRANSACTION_TYPE'].isin(excluded_types))
       ]

       # Step 5: Group by STORE_NAME and sum AMOUNT
       asp_pending_sum = asp_pending_data.groupby('STORE_NAME')['AMOUNT'].sum().reset_index()
       asp_pending_sum.columns = ['Store_amend', 'Asp_Pending']

       # ✅ Step 6: Remove all existing Asp_Pending columns to avoid duplicates
       store_summary1 = store_summary1.loc[:, ~store_summary1.columns.str.contains('Asp_Pending', case=False)]

       # Step 7: Merge clean Asp_Pending column
       store_summary1 = store_summary1.merge(asp_pending_sum, on='Store_amend', how='left')
       store_summary1['Asp_Pending'] = store_summary1['Asp_Pending'].fillna(0)

       # Step 8: Move Asp_Pending column after Reversals
       cols = list(store_summary1.columns)
       cols.remove('Asp_Pending')
       insert_at = cols.index('Reversals') + 1 if 'Reversals' in cols else len(cols)
       cols.insert(insert_at, 'Asp_Pending')
       store_summary1 = store_summary1[cols]

       # Step 9: Recompute TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Final Preview

       store_summary1.tail()


       # In[ ]:


       # Step 1: Ensure CREDIT_AMOUNT is numeric
       safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')

       # Step 2: Get the mode of Ref1 in safaricom
       most_common_ref1 = safaricom['Ref1'].mode()[0]

       # Step 3: Filter safaricom data based on the six conditions
       unsync_data = safaricom[
           (safaricom['ACCOUNT_TYPE_NAME'] == 'Merchant Account') &
           (safaricom['code_check'] == 'XX') &
           (safaricom['Ref1'] == most_common_ref1)
       ]

       # Step 4: Group by Store_amend and sum CREDIT_AMOUNT
       unsync_sum = unsync_data.groupby('Store_amend')['CREDIT_AMOUNT'].sum().reset_index()
       unsync_sum.columns = ['Store_amend', 'unsync']

       # Step 5: Merge into store_summary1
       store_summary1 = store_summary1.merge(unsync_sum, on='Store_amend', how='left')
       store_summary1['unsync'] = store_summary1['unsync'].fillna(0)

       # Step 6: (Optional) Move column if needed
       cols = list(store_summary1.columns)
       if 'unsync' in cols:
           cols.remove('unsync')
           insert_at = cols.index('saf_paid') + 1 if 'saf_paid' in cols else len(cols)
           cols.insert(insert_at, 'unsync')
           store_summary1 = store_summary1[cols]

       # Step 7: Recalculate TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Final output
       store_summary1.tail()



       # In[ ]:


       # ✅ Final preview
       store_summary1.tail()


       # In[ ]:


       # Step 1: Ensure all required columns are numeric
       cols_to_check = ['unutilized_txn', 'unsync', 'Asp_Pending']
       for col in cols_to_check:
           store_summary1[col] = pd.to_numeric(store_summary1[col], errors='coerce').fillna(0)

       # Step 2: Compute Reversal Charges
       store_summary1['Reversal Charges'] = store_summary1['unutilized_txn'] - store_summary1['unsync'] - store_summary1['Asp_Pending']

       # Step 3: Reorder - place Reversal Charges after Asp_Pending
       cols = list(store_summary1.columns)
       if 'Reversal Charges' in cols and 'Asp_Pending' in cols:
           cols.remove('Reversal Charges')
           insert_at = cols.index('Asp_Pending') + 1
           cols.insert(insert_at, 'Reversal Charges')
           store_summary1 = store_summary1[cols]

       # ✅ Final Preview
       store_summary1.tail()


       # In[ ]:


       # Step 1: Ensure all required columns are numeric
       required_cols = ['unutilized_txn', 'unsync', 'Asp_Pending', 'Reversal Charges']
       for col in required_cols:
           store_summary1[col] = pd.to_numeric(store_summary1[col], errors='coerce').fillna(0)

       # Step 2: Calculate Variance
       store_summary1['Variance'] = (
           store_summary1['unutilized_txn']
           - store_summary1['unsync']
           - store_summary1['Asp_Pending']
           - store_summary1['Reversal Charges']
       )

       # Step 3: Reorder 'Variance' to appear after 'Reversal Charges'
       cols = list(store_summary1.columns)
       if 'Variance' in cols and 'Reversal Charges' in cols:
           cols.remove('Variance')
           insert_at = cols.index('Reversal Charges') + 1
           cols.insert(insert_at, 'Variance')
           store_summary1 = store_summary1[cols]

       # Step 4: Recompute TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # ✅ Final Preview
       store_summary1.tail()


       # ###Compute Variances for the day

       # In[ ]:


       # Step 1: Ensure all required columns are numeric
       required_cols = ['unutilized_txn', 'unsync', 'Asp_Pending', 'Reversal Charges']
       for col in required_cols:
           store_summary1[col] = pd.to_numeric(store_summary1[col], errors='coerce').fillna(0)

       # Step 2: Calculate Variance
       store_summary1['Variance'] = (
           store_summary1['unutilized_txn']
           - store_summary1['unsync']
           - store_summary1['Asp_Pending']
           - store_summary1['Reversal Charges']
       )

       # Step 3: Reorder 'Variance' to appear after 'Reversal Charges'
       cols = list(store_summary1.columns)
       if 'Variance' in cols and 'Reversal Charges' in cols:
           cols.remove('Variance')
           insert_at = cols.index('Reversal Charges') + 1
           cols.insert(insert_at, 'Variance')
           store_summary1 = store_summary1[cols]

       # Step 4: Recompute TOTAL row
       store_summary1 = store_summary1[store_summary1['Store_amend'] != 'TOTAL']
       numeric_cols = store_summary1.select_dtypes(include='number').columns
       totals = store_summary1[numeric_cols].sum().to_dict()
       totals_row = {col: '' for col in store_summary1.columns}
       totals_row.update(totals)
       totals_row['Store_amend'] = 'TOTAL'
       store_summary1 = pd.concat([store_summary1, pd.DataFrame([totals_row])], ignore_index=True)

       # Step 5: Reorder columns to desired format
       desired_order = [
           'Store_amend', 'Bank_Transfer', 'Charges', 'Prev_day',
           'Asp_Utilized', 'saf_paid', 'unutilized_txn', 'unsync',
           'Asp_Pending', 'Reversals', 'Reversal Charges'
       ]
       store_summary1 = store_summary1[desired_order]

       # Final Preview
       store_summary1.tail()


       # #3.Csv Reports Generation

       # ##Generate list of Reversals done

       # In[ ]:


       import pandas as pd
       # [REMOVED FOR DEPLOYMENT] from google.colab import files

       # Ensure the DEBIT_AMOUNT column is numeric
       safaricom['DEBIT_AMOUNT'] = pd.to_numeric(safaricom['DEBIT_AMOUNT'], errors='coerce')

       # Filter for reversal rows with a valid LINKED_TRANSACTION_ID and non-zero DEBIT_AMOUNT
       reversal_rows = safaricom[
           safaricom['LINKED_TRANSACTION_ID'].notna() &
           (safaricom['DEBIT_AMOUNT'] > 0)
       ]

       # Select the required columns in the desired order
       daily_reversals = reversal_rows[['Store_amend', 'TRANSACTION_PARTY_DETAILS', 'DEBIT_AMOUNT', 'LINKED_TRANSACTION_ID']]

       # Export to CSV
       daily_reversals.to_csv('daily_reversals.csv', index=False)

       # Trigger auto-download (Colab)
       if os.path.exists('daily_reversals.csv'):


       # ##Generate previous day transactions utilized

       # In[ ]:


       import pandas as pd
       # [REMOVED FOR DEPLOYMENT] from google.colab import files

       # Step 0: Clean up TRANSACTION_TYPE for prioritization
       aspire['TRANSACTION_TYPE'] = aspire['TRANSACTION_TYPE'].fillna('').astype(str).str.strip()
       preferred_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       # Step 0.1: Identify duplicated TRANSACTION_IDs
       dupes = aspire[aspire.duplicated('TRANSACTION_ID', keep=False)]

       # Step 0.2: Add priority column
       dupes['priority'] = dupes['TRANSACTION_TYPE'].apply(
           lambda x: 0 if x in preferred_types else (1 if x else 2)
       )

       # Step 0.3: Keep one best row per TRANSACTION_ID
       cleaned_dupes = (
           dupes.sort_values(by='priority')
                .drop_duplicates(subset='TRANSACTION_ID', keep='first')
                .drop(columns='priority')
       )

       # Step 0.4: Combine cleaned dupes with non-dupes
       aspire_unique = aspire[~aspire['TRANSACTION_ID'].isin(dupes['TRANSACTION_ID'])]
       aspire = pd.concat([aspire_unique, cleaned_dupes], ignore_index=True)

       # Step 1: Get the most frequent Ref1 in safaricom
       most_common_ref1 = safaricom['Ref1'].mode()[0]

       # Step 2: Filter for valid rows
       valid_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']
       filtered_aspire = aspire[
           (aspire['Ref1'] != most_common_ref1) &
           (aspire['TRANSACTION_TYPE'].isin(valid_types))
       ]

       # Step 3: Select required columns for export
       prev_day_utilized = filtered_aspire[['STORE_NAME', 'TRANSACTION_ID', 'AMOUNT', 'SYSTEM_ENTRY_DATE', 'TRANSACTION_TYPE']]

       # Step 4: Export to CSV and auto-download
       prev_day_utilized.to_csv('Prev_day_utilized.csv', index=False)
       if os.path.exists('daily_reversals.csv'):


       # ##Generate unutilized transactions

       # In[ ]:


       # === Step 1: Clean and prepare 'aspire' transactions (Asp_Pending) ===
       aspire['AMOUNT'] = pd.to_numeric(aspire['AMOUNT'], errors='coerce')
       aspire['SYSTEM_ENTRY_DATE'] = pd.to_datetime(aspire['SYSTEM_ENTRY_DATE'], errors='coerce')

       most_common_ref1 = aspire['Ref1'].mode()[0]
       excluded_types = ['POS CASH SALE', 'DEPOSIT RECEIVED']

       asp_pending = aspire[
           (aspire['Ref1'] == most_common_ref1) &
           (~aspire['TRANSACTION_TYPE'].isin(excluded_types))
       ].copy()

       asp_pending['VENDOR_TIME'] = asp_pending['SYSTEM_ENTRY_DATE'].dt.strftime('%H:%M')
       asp_pending['VENDOR_DAY'] = asp_pending['SYSTEM_ENTRY_DATE'].dt.strftime('%d/%m/%Y')

       asp_export = asp_pending[['VENDOR_TIME', 'STORE_NAME', 'TRANSACTION_ID', 'VENDOR_DAY', 'AMOUNT']].copy()

       # === Step 2: Prepare 'safaricom' unsynced transactions ===
       safaricom['CREDIT_AMOUNT'] = pd.to_numeric(safaricom['CREDIT_AMOUNT'], errors='coerce')
       safaricom['START_TIMESTAMP'] = pd.to_datetime(safaricom['START_TIMESTAMP'], errors='coerce')

       most_common_ref1_saf = safaricom['Ref1'].mode()[0]

       unsync = safaricom[
           (safaricom['ACCOUNT_TYPE_NAME'] == 'Merchant Account') &
           (safaricom['code_check'] == 'XX') &
           (safaricom['Ref1'] == most_common_ref1_saf) &
           (safaricom['LINKED_TRANSACTION_ID'].isna())  # ✅ Exclude if linked
       ].copy()

       unsync['VENDOR_TIME'] = unsync['START_TIMESTAMP'].dt.strftime('%H:%M')
       unsync['VENDOR_DAY'] = unsync['START_TIMESTAMP'].dt.strftime('%d/%m/%Y')

       unsync_export = unsync[['VENDOR_TIME', 'Store_amend', 'RECEIPT_NUMBER', 'VENDOR_DAY', 'CREDIT_AMOUNT']].copy()
       unsync_export.columns = ['VENDOR_TIME', 'STORE_NAME', 'TRANSACTION_ID', 'VENDOR_DAY', 'AMOUNT']
       unsync_export = unsync_export[unsync_export['AMOUNT'] > 1]

       # === Step 3: Map STORE_NAME using the 'key' DataFrame ===
       store_name_map = dict(zip(key.iloc[:, 0], key.iloc[:, 1]))
       asp_export['STORE_NAME'] = asp_export['STORE_NAME'].map(store_name_map).fillna(asp_export['STORE_NAME'])
       unsync_export['STORE_NAME'] = unsync_export['STORE_NAME'].map(store_name_map).fillna(unsync_export['STORE_NAME'])

       # === Step 4: Combine both DataFrames ===
       cashed_out = pd.concat([asp_export, unsync_export], ignore_index=True)

       # === Step 5: Sort and format ===
       cashed_out = cashed_out.sort_values(by=['STORE_NAME', 'AMOUNT'], ascending=[True, False])

       # === Step 6: Export CSV ===
       cashed_out.to_csv("Cashed_out.csv", index=False)
       if os.path.exists('daily_reversals.csv'):


       # ##Generate the Daily Summary

       # In[ ]:


       # === Load file and extract date from filename ===
       filename = '/content/Mpesa_852182_20250611235959.csv'
       safaricom = pd.read_csv(filename)

       match = re.search(r'(\d{8})', filename)
       if match:
           raw_date = match.group(1)
           formatted_date = datetime.strptime(raw_date, "%Y%m%d").strftime("%d-%b-%y")
       else:
           formatted_date = "Date Unknown"

       # === Prepare data (store_summary1 should exist) ===
       df = store_summary1.copy()

       # Remove totals row and insert serial number
       df_data_only = df[~df['Store_amend'].str.lower().str.contains("total", na=False)].copy()
       df_data_only.insert(0, "S/No", range(1, len(df_data_only) + 1))

       # === Add totals row ===
       totals_row = ["Total"]
       for col in df_data_only.columns[1:]:  # Skip S/No
           if pd.api.types.is_numeric_dtype(df_data_only[col]):
               totals_row.append(df_data_only[col].sum())
           else:
               totals_row.append("")

       df_data_only.loc[len(df_data_only)] = totals_row

       # === Create workbook ===
       wb = Workbook()
       ws = wb.active
       ws.title = "Reconciliation"

       # Add title and date
       ws.append(["Mpesa Reconciliation"])
       ws.append([formatted_date])

       # Add column headers
       ws.append(df_data_only.columns.tolist())

       # Add data rows
       for row in df_data_only.itertuples(index=False):
           ws.append(row)

       # === Style: Bold Title and Headers ===
       ws["A1"].font = Font(bold=True, size=14)
       ws["A2"].font = Font(italic=True)

       header_row = 3
       for col in range(1, len(df_data_only.columns) + 1):
           cell = ws.cell(row=header_row, column=col)
           cell.font = Font(bold=True)

       # === Freeze at B4 (first data row) ===
       ws.freeze_panes = "B4"

       # === AutoFilter ===
       ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{header_row}"

       # === Save and Download ===
       excel_file = "store_summary1.xlsx"
       wb.save(excel_file)
       if os.path.exists('daily_reversals.csv'):

               # 🔄 Prepare Excel workbook with multiple sheets
               output_excel = BytesIO()
               with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
                   if 'daily_reversals' in locals():
                       daily_reversals.to_excel(writer, sheet_name='Daily_Reversals', index=False)
                   if 'utilized' in locals():
                       utilized.to_excel(writer, sheet_name='Utilized', index=False)
                   if 'not_utilized' in locals():
                       not_utilized.to_excel(writer, sheet_name='Not_Utilized', index=False)
                   if 'final_output' in locals():
                       final_output.to_excel(writer, sheet_name='Final_Output', index=False)
               output_excel.seek(0)

               st.download_button(
                   label="📥 Download All Reports (Excel Workbook)",
                   data=output_excel,
                   file_name="mpesa_reconciliation_report.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   key="btn_all_reports"
               )


        if 'final_output' in locals():
            st.success("✅ Processing complete. Final Output:")
            st.dataframe(final_output)
